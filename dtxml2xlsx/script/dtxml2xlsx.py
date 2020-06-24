from os import path, remove
from sys import argv
import xml.etree.ElementTree as ET
import openpyxl

def doConvert(filename):
    tree = ET.parse(filename)

    root = tree.getroot()
    wbname = root.attrib['Schema'] + '.xlsx'

    if path.exists(wbname) and path.isfile(wbname):
        remove(wbname)

    wb = openpyxl.Workbook()

    for sheet in root.iter('Sheet'):
        st = wb.create_sheet(index=len(wb.sheetnames)-1, title=sheet.attrib['Name'])
        key_id = {}
        row_num = 1
        for cols in sheet.iter('Columns'):
            col_num = 1
            for col in cols.iter('Column'):
                key = col.attrib['Name']
                key_id[key] = col_num
                st.cell(row_num, col_num, value=key)
                col_num = col_num + 1
        for row in sheet.iter('Row'):
            row_num = row_num + 1
            for cell in row.iter('Cell'):
                key = cell.attrib['Name']
                value = cell.text
                st.cell(row_num, key_id[key], value=value)

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    wb.save(wbname)

if __name__ == '__main__':
    if len(argv) > 1:
        doConvert(argv[1])
