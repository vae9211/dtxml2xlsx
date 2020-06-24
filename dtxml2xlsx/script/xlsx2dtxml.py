from os import path, rename
from sys import argv
import xml.etree.ElementTree as ET
import openpyxl

def doConvert(wbname):
    pathinfo = path.split(wbname)
    fileinfo = path.splitext(pathinfo[1])

    if fileinfo[1] != '.xlsx':
        print('不是xlsx文件', wbname)
        return

    wb = openpyxl.load_workbook(wbname)
    filename = wbname.replace('xlsx', 'dtxml')

    if path.exists(filename) and path.isfile(filename):
        pass
    else:
        print('找不到dtxml文件', filename)
        return

    tree = ET.parse(filename)
    root = tree.getroot()

    for sheet in root.iter('Sheet'):
        st = wb[sheet.attrib['Name']]
        key_id = {}
        row_num = 1
        for cols in sheet.iter('Columns'):
            col_num = 1
            for col in cols.iter('Column'):
                c = st.cell(row_num, col_num)
                key_id[c.value] = col_num
                col_num = col_num + 1
        for row in sheet.iter('Row'):
            row_num = row_num + 1
            for cell in row.iter('Cell'):
                c = st.cell(row_num, key_id[cell.attrib['Name']])
                cell.text = str(c.value)

    rename(filename, filename + '.bak')
    tree.write(filename, encoding='UTF-8')

if __name__ == '__main__':
    if len(argv) > 1:
        doConvert(argv[1])
